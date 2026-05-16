"""
Reads the Excel spreadsheet and generates a browser-ready import script.
Deduplicates students, normalizes data, and maps tutors.
"""
import openpyxl
import json
import re

SHEET_TO_TUTOR = {
    'Rob': 'Rob Daglish',
    'Cello': 'Naomi Harmer',
    'Lana': 'Lana Law',
    'Drums': 'Cameron Finlay',
    'Flute': 'Susan Dollin',
    'Guitar': 'Motoi Shibusawa',
    'Piano - Christine': 'Christine Rudd',
    'Piano - Lynley': 'Lynley Fuglestad',
    'Piano - Marina': 'Marina Oulton',
    'Brass': 'Julian Marchant',
    'Violin': 'Julie Pettitt',
    'Vocals': 'Elizabeth Macfarlane',
}

# Default instrument for each sheet (used when instrument column is empty)
SHEET_DEFAULT_INSTRUMENT = {
    'Rob': 'Bass',
    'Cello': 'Cello',
    'Lana': 'Saxophone',
    'Drums': 'Drums',
    'Flute': 'Flute',
    'Guitar': 'Guitar',
    'Piano - Christine': 'Piano',
    'Piano - Lynley': 'Piano',
    'Piano - Marina': 'Piano',
    'Brass': 'Trumpet',
    'Violin': 'Violin',
    'Vocals': 'Vocals',
}

# Column name normalization - map various header names to standard keys
COLUMN_MAP = {
    "child's full name": 'name',
    "student name": 'name',
    "child's email address (year 9 and above only)": 'email',
    "parent full name": 'parentName',
    "parent email address": 'parentEmail',
    "parent phone number": 'parentPhone',
    "year level in 2026": 'year',
    "is your child taking music as an option subject in 2026?": 'musicOption',
    "which instrument would you like lessons for?": 'instrument',
    "what instrument do they play?": 'instrumentExisting',
    "do you have a preference in tutor": 'tutorPreference',
    "please indicate your preferred style (tick all that apply)": 'style',
    "please let us know about any previous experience and current grade": 'experience',
    "if your child plays any other instruments, please let us know. please state instrument and grade.1": 'otherInstruments',
    "if your child plays any other instruments, please let us know. please state instrument and grade.": 'otherInstrumentsAlt',
    "please give the name of the tutor": 'existingTutor',
    "tutor's email address (or phone number if email is unknown)": 'existingTutorEmail',
    "what grade is your child currently playing at": 'currentGrade',
}

def normalize_header(h):
    """Map a header to a standard key."""
    if not h:
        return None
    h_clean = h.strip().lower().replace('\xa0', ' ').replace('\n', ' ')
    # Remove trailing numbers from duplicate columns
    h_clean = re.sub(r'\d+$', '', h_clean).strip()
    # Try exact match first
    for pattern, key in COLUMN_MAP.items():
        if h_clean.startswith(pattern[:40]):
            return key
    # Check for funding column (long name)
    if 'please select one of the following' in h_clean and 'funded' in h_clean:
        return 'fundingType'
    if 'please select one of the following' in h_clean:
        return 'fundingType'
    return None

def normalize_funding(val):
    """Normalize funding status."""
    if not val:
        return None
    val = val.strip().lower()
    if 'funded' in val:
        return 'Funded Requested'
    if 'private' in val:
        return 'Private'
    return val

def normalize_instrument(val):
    """Normalize instrument name."""
    if not val:
        return None
    val = val.strip()
    # Remove "(Y9-13 only)" suffix from Vocals
    val = re.sub(r'\s*\(Y\d+-\d+ only\)', '', val)
    return val

def clean_str(val):
    """Clean a string value."""
    if val is None:
        return ''
    val = str(val).strip()
    # Replace mojibake characters
    val = val.replace('\u2019', "'").replace('\u2018', "'")
    val = val.replace('\u201c', '"').replace('\u201d', '"')
    val = val.replace('\ufffd', "'")
    return val

def parse_spreadsheet(filepath):
    wb = openpyxl.load_workbook(filepath)
    all_rows = []

    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEET_TO_TUTOR:
            print(f"Skipping unknown sheet: {sheet_name}")
            continue

        ws = wb[sheet_name]
        tutor_name = SHEET_TO_TUTOR[sheet_name]

        # Parse headers from row 1
        headers = []
        for cell in ws[1]:
            headers.append(normalize_header(cell.value))

        # Parse data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = {}
            for col_idx, val in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    key = headers[col_idx]
                    if val is not None:
                        # Handle multiple columns mapping to same key (take first non-null)
                        if key == 'fundingType':
                            if key not in data or not data[key]:
                                data[key] = normalize_funding(str(val))
                        elif key == 'otherInstruments' or key == 'otherInstrumentsAlt':
                            data.setdefault('otherInstruments', clean_str(val))
                        else:
                            data.setdefault(key, val)

            # Get name
            name = clean_str(data.get('name', ''))
            if not name:
                continue

            # Get instrument - from 'instrument' column or 'instrumentExisting' or default from sheet
            instrument = normalize_instrument(clean_str(data.get('instrument', '')))
            if not instrument:
                instrument = normalize_instrument(clean_str(data.get('instrumentExisting', '')))
            if not instrument:
                instrument = SHEET_DEFAULT_INSTRUMENT.get(sheet_name)
                if instrument:
                    print(f"  Defaulting {name} in {sheet_name} to {instrument}")
                else:
                    print(f"  Skipping {name} in {sheet_name} - no instrument and no default")
                    continue

            # Determine funding
            funding = data.get('fundingType')
            if not funding:
                funding = None

            # Determine actual tutor - special case for flute students on Piano-Lynley sheet
            actual_tutor = tutor_name
            if instrument == 'Flute' and tutor_name != 'Susan Dollin':
                actual_tutor = 'Susan Dollin'

            year_val = data.get('year')
            if year_val is not None:
                try:
                    year_val = int(year_val)
                except (ValueError, TypeError):
                    year_val = None

            row_data = {
                'name': name,
                'email': clean_str(data.get('email', '')),
                'parentName': clean_str(data.get('parentName', '')),
                'parentEmail': clean_str(data.get('parentEmail', '')),
                'parentPhone': clean_str(data.get('parentPhone', '')),
                'year': year_val,
                'musicOption': clean_str(data.get('musicOption', '')),
                'instrument': instrument,
                'funded': funding,
                'experience': clean_str(data.get('experience', '')),
                'otherInstruments': clean_str(data.get('otherInstruments', '')),
                'style': clean_str(data.get('style', '')),
                'tutorName': actual_tutor,
                'sheet': sheet_name,
            }
            all_rows.append(row_data)

    return all_rows

def deduplicate_students(rows):
    """Deduplicate students by normalized name + parent email."""
    students = {}
    for row in rows:
        key = f"{row['name'].lower().strip()}|{row['parentEmail'].lower().strip()}"
        if key not in students:
            students[key] = {
                'name': row['name'],
                'email': row['email'],
                'year': row['year'],
                'parentName': row['parentName'],
                'parentEmail': row['parentEmail'],
                'parentPhone': row['parentPhone'],
                'instruments': [row['instrument']],
                'status': 'active',
                'musicOption': row['musicOption'] == 'Yes',
            }
        else:
            # Update with non-empty values and add instrument
            s = students[key]
            if row['email'] and not s['email']:
                s['email'] = row['email']
            if row['instrument'] not in s['instruments']:
                s['instruments'].append(row['instrument'])
    return students

def deduplicate_lessons(rows):
    """Deduplicate lessons by student name + parent email + instrument + tutor."""
    seen = set()
    unique = []
    for row in rows:
        key = f"{row['name'].lower().strip()}|{row['parentEmail'].lower().strip()}|{row['instrument'].lower()}|{row['tutorName'].lower()}"
        if key not in seen:
            seen.add(key)
            unique.append(row)
        else:
            print(f"  Dedup lesson: {row['name']} - {row['instrument']} ({row['tutorName']}) from {row['sheet']}")
    return unique

def generate_import_script(students, lessons):
    """Generate the browser-ready import script."""

    students_json = json.dumps(list(students.values()), indent=4, ensure_ascii=False)
    lessons_json = json.dumps([{
        'studentKey': f"{r['name'].lower().strip()}|{r['parentEmail'].lower().strip()}",
        'studentName': r['name'],
        'tutorName': r['tutorName'],
        'instrument': r['instrument'],
        'funded': r['funded'] is not None and 'funded' in r['funded'].lower() if r['funded'] else False,
        'parentName': r['parentName'],
        'parentEmail': r['parentEmail'],
        'parentPhone': r['parentPhone'],
        'notes': '\n'.join(filter(None, [
            f"Experience: {r['experience']}" if r['experience'] else '',
            f"Other instruments: {r['otherInstruments']}" if r['otherInstruments'] else '',
            f"Style preference: {r['style']}" if r['style'] else '',
        ])),
    } for r in lessons], indent=4, ensure_ascii=False)

    script = f"""/**
 * MGS Arts Portal - Auto-generated Import Script
 * Generated from: Copy of Itinerant Music Lessons 2026.xlsx
 *
 * Run in the browser console while logged into the admin portal:
 *   1. Copy-paste this entire script
 *   2. Call: await runImport()
 *   3. Reload the page
 */

const STUDENTS = {students_json};

const LESSONS = {lessons_json};

async function runImport() {{
    console.log('🎵 MGS Arts Portal - Lesson Import');
    console.log('==================================');

    // Step 0: Load existing tutors
    console.log('\\n📋 Loading existing tutors...');
    const tutors = await DatabaseService.getTutors();
    if (!tutors || tutors.length === 0) {{
        console.error('❌ No tutors found in Firestore. Please add tutors first.');
        return;
    }}

    const tutorLookup = {{}};
    tutors.forEach(t => {{
        tutorLookup[t.name.toLowerCase().trim()] = t;
    }});
    console.log(`Found ${{tutors.length}} tutors:`, tutors.map(t => t.name).join(', '));

    // Validate tutor mappings
    const requiredTutors = [...new Set(LESSONS.map(l => l.tutorName))];
    const missing = requiredTutors.filter(t => !tutorLookup[t.toLowerCase().trim()]);
    if (missing.length > 0) {{
        console.error('❌ Missing tutors:', missing);
        console.log('Available:', Object.keys(tutorLookup));
        return;
    }}
    console.log('✅ All tutor mappings validated');

    // Step 1: Create students
    console.log(`\\n👨‍🎓 Creating ${{STUDENTS.length}} students...`);
    const studentIdMap = {{}};
    let studentOk = 0, studentErr = 0;

    for (const student of STUDENTS) {{
        const key = `${{student.name.toLowerCase().trim()}}|${{(student.parentEmail || '').toLowerCase().trim()}}`;
        const result = await DatabaseService.addStudent(student);
        if (result.success) {{
            studentIdMap[key] = result.id;
            studentOk++;
            if (studentOk % 20 === 0) console.log(`  ... ${{studentOk}} students created`);
        }} else {{
            console.error(`  Failed: ${{student.name}}`, result.error);
            studentErr++;
        }}
    }}
    console.log(`✅ Students: ${{studentOk}} created, ${{studentErr}} errors`);

    // Step 2: Create lessons
    console.log(`\\n🎶 Creating ${{LESSONS.length}} lessons...`);
    let lessonOk = 0, lessonErr = 0;

    for (const l of LESSONS) {{
        const studentId = studentIdMap[l.studentKey];
        const tutor = tutorLookup[l.tutorName.toLowerCase().trim()];

        if (!studentId) {{
            console.error(`  No student ID for: ${{l.studentName}} (key: ${{l.studentKey}})`);
            lessonErr++;
            continue;
        }}

        const lesson = {{
            studentId: studentId,
            studentName: l.studentName,
            tutorId: tutor.id,
            tutorName: tutor.name,
            instrument: l.instrument,
            day: '',
            time: '',
            status: 'active',
            funded: l.funded,
            parentName: l.parentName,
            parentEmail: l.parentEmail,
            parentPhone: l.parentPhone,
            notes: l.notes
        }};

        const result = await DatabaseService.addLesson(lesson);
        if (result.success) {{
            lessonOk++;
            if (lessonOk % 20 === 0) console.log(`  ... ${{lessonOk}} lessons created`);
        }} else {{
            console.error(`  Failed: ${{l.studentName}} - ${{l.instrument}}`, result.error);
            lessonErr++;
        }}
    }}
    console.log(`✅ Lessons: ${{lessonOk}} created, ${{lessonErr}} errors`);

    // Summary
    console.log('\\n==================================');
    console.log('🎉 Import complete!');
    console.log(`   Students: ${{studentOk}}`);
    console.log(`   Lessons:  ${{lessonOk}}`);
    console.log(`   Errors:   ${{studentErr + lessonErr}}`);
    console.log('\\n💡 Reload the page to see the imported data.');
}}
"""
    return script

if __name__ == '__main__':
    filepath = 'D:/Sidequest Digital/Dev Projects/Clients/MGSArtsPortal/Copy of Itinerant Music Lessons 2026.xlsx'
    print("Reading spreadsheet...")
    rows = parse_spreadsheet(filepath)
    print(f"Parsed {len(rows)} total rows")

    print("\nDeduplicating students...")
    students = deduplicate_students(rows)
    print(f"Found {len(students)} unique students")

    print("\nDeduplicating lessons...")
    lessons = deduplicate_lessons(rows)
    print(f"Found {len(lessons)} unique lessons")

    print("\nGenerating import script...")
    script = generate_import_script(students, lessons)

    output_path = 'D:/Sidequest Digital/Dev Projects/Clients/MGSArtsPortal/import-lessons.js'
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(script)

    print(f"\nImport script written to: {output_path}")
    print(f"   Students: {len(students)}")
    print(f"   Lessons:  {len(lessons)}")
    print("\nTo use: open the admin portal in a browser, open the console, paste the script, and run: await runImport()")
